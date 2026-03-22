# gerar_excel.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# 1. DADOS
# ─────────────────────────────────────────
df       = pd.read_csv("output/dados_completos.csv")
sessoes  = pd.read_csv("output/sessoes.csv")
veic_mot = pd.read_csv("output/motoristas_veiculos.csv")

sessoes["Inicio"]      = pd.to_datetime(sessoes["Inicio"])
sessoes["Fim"]         = pd.to_datetime(sessoes["Fim"])
sessoes["Duracao_min"] = sessoes["Duracao_min"].round(1)

# ─────────────────────────────────────────
# 2. ESTILOS
# ─────────────────────────────────────────
AZUL_ESC  = "0D2B5E"
AZUL_CLR  = "EAF1FB"
VERDE_ESC = "1B5E20"
VERDE_CLR = "E8F5E9"
CINZA     = "F2F4F7"
BRANCO    = "FFFFFF"

def header_style(ws, row, cols, cor_fundo=AZUL_ESC, cor_texto=BRANCO):
    fill = PatternFill("solid", fgColor=cor_fundo)
    font = Font(bold=True, color=cor_texto, name="Arial", size=10)
    alin = Alignment(horizontal="center", vertical="center")
    bord = borda_fina()
    for col in range(1, cols+1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = alin
        c.border = bord

def borda_fina():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def celula(ws, row, col, valor, bold=False, cor_fundo=None, alinhamento="left", numero=False):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(name="Arial", size=10, bold=bold)
    c.alignment = Alignment(horizontal=alinhamento, vertical="center")
    c.border = borda_fina()
    if cor_fundo:
        c.fill = PatternFill("solid", fgColor=cor_fundo)
    if numero:
        c.number_format = '#,##0.0'
    return c

def titulo_aba(ws, texto, subtexto=None):
    ws.row_dimensions[1].height = 36
    t = ws.cell(row=1, column=1, value=texto)
    t.font = Font(name="Arial", size=16, bold=True, color=AZUL_ESC)
    t.alignment = Alignment(vertical="center")
    if subtexto:
        ws.row_dimensions[2].height = 20
        s = ws.cell(row=2, column=1, value=subtexto)
        s.font = Font(name="Arial", size=10, color="666666", italic=True)

# ─────────────────────────────────────────
# 3. WORKBOOK
# ─────────────────────────────────────────
wb = Workbook()

# ── ABA 1: DADOS COMPLETOS ────────────────
ws1 = wb.active
ws1.title = "Dados Completos"
titulo_aba(ws1, "Dados Completos — Rastreamento GPS",
           "Todos os registros de posição com status de geofence")

headers = ["Veículo", "Data", "Dia da Semana", "Hora",
           "Latitude", "Longitude", "Motorista", "Status"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
header_style(ws1, 4, len(headers))

for ri, row in enumerate(df.itertuples(index=False), 5):
    bg = VERDE_CLR if row.Status == "Dentro" else (CINZA if ri % 2 == 0 else BRANCO)
    for ci, val in enumerate(row, 1):
        c = celula(ws1, ri, ci, val, cor_fundo=bg,
                   alinhamento="center" if ci > 1 else "left")

larguras1 = [10, 12, 14, 10, 14, 14, 20, 10]
for i, w in enumerate(larguras1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A5"
ws1.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
print("  ✓ Aba 'Dados Completos' criada")

# ── ABA 2: SESSÕES ────────────────────────
ws2 = wb.create_sheet("Sessões na Base")
titulo_aba(ws2, "Sessões dentro da Base Operacional",
           "Períodos contínuos com veículo dentro da cerca eletrônica")

sess_dentro = sessoes[sessoes["Status"] == "Dentro"].copy()
sess_dentro["Inicio_str"] = sess_dentro["Inicio"].dt.strftime("%d/%m/%Y %H:%M")
sess_dentro["Fim_str"]    = sess_dentro["Fim"].dt.strftime("%d/%m/%Y %H:%M")
sess_dentro["Data"]       = sess_dentro["Inicio"].dt.strftime("%d/%m/%Y")
sess_dentro["Dia"]        = sess_dentro["Inicio"].dt.day_name().map({
    "Monday":"Segunda","Tuesday":"Terça","Wednesday":"Quarta",
    "Thursday":"Quinta","Friday":"Sexta"
})

h2 = ["#", "Veículo", "Data", "Dia da Semana", "Entrada", "Saída",
      "Duração (min)", "Motorista"]
for i, h in enumerate(h2, 1):
    ws2.cell(row=4, column=i, value=h)
header_style(ws2, 4, len(h2), cor_fundo="1B5E20")

for ri, row in enumerate(sess_dentro.itertuples(index=False), 5):
    bg = VERDE_CLR if ri % 2 == 0 else BRANCO
    vals = [ri-4, row.Veículo, row.Data, row.Dia,
            row.Inicio_str, row.Fim_str, row.Duracao_min, row.Motorista]
    for ci, val in enumerate(vals, 1):
        celula(ws2, ri, ci, val, cor_fundo=bg,
               alinhamento="center" if ci != 8 else "left",
               numero=(ci == 7))

larguras2 = [4, 10, 12, 14, 16, 16, 14, 20]
for i, w in enumerate(larguras2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A5"
print("  ✓ Aba 'Sessões na Base' criada")

# ── ABA 3: RESUMO VEÍCULOS ────────────────
ws3 = wb.create_sheet("Resumo Veículos")
titulo_aba(ws3, "Resumo por Veículo",
           "Tempo total dentro e fora da Base Operacional na semana")

res_vei = sessoes.groupby(["Veículo","Status"])["Duracao_min"].sum().unstack(fill_value=0)
res_vei.columns.name = None
if "Dentro" not in res_vei.columns: res_vei["Dentro"] = 0
if "Fora"   not in res_vei.columns: res_vei["Fora"]   = 0
res_vei = res_vei.reset_index()
res_vei["Dentro_h"] = (res_vei["Dentro"]/60).round(2)
res_vei["Fora_h"]   = (res_vei["Fora"]/60).round(2)
res_vei["Pct_Dentro"] = (res_vei["Dentro"] / (res_vei["Dentro"]+res_vei["Fora"]) * 100).round(1)
regs = df.groupby("Veículo").size().reset_index(name="Registros")
res_vei = res_vei.merge(regs, on="Veículo")

h3 = ["Veículo","Dentro (min)","Dentro (h)","Fora (min)","Fora (h)","% na Base","Registros GPS"]
for i, h in enumerate(h3, 1):
    ws3.cell(row=4, column=i, value=h)
header_style(ws3, 4, len(h3))

for ri, row in enumerate(res_vei.itertuples(index=False), 5):
    bg = VERDE_CLR if row.Dentro > 0 else (CINZA if ri % 2 == 0 else BRANCO)
    vals = [row.Veículo, round(row.Dentro,1), row.Dentro_h,
            round(row.Fora,1), row.Fora_h,
            f"{row.Pct_Dentro}%", row.Registros]
    for ci, val in enumerate(vals, 1):
        bold = (row.Dentro > 0 and ci in [1,2,3,6])
        celula(ws3, ri, ci, val, bold=bold, cor_fundo=bg, alinhamento="center")

larguras3 = [10, 14, 12, 12, 10, 12, 14]
for i, w in enumerate(larguras3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = "A5"
print("  ✓ Aba 'Resumo Veículos' criada")

# ── ABA 4: RESUMO MOTORISTAS ──────────────
ws4 = wb.create_sheet("Resumo Motoristas")
titulo_aba(ws4, "Resumo por Motorista",
           "Tempo operado dentro e fora da Base por cada motorista")

res_mot = sessoes.groupby(["Motorista","Status"])["Duracao_min"].sum().unstack(fill_value=0)
res_mot.columns.name = None
if "Dentro" not in res_mot.columns: res_mot["Dentro"] = 0
if "Fora"   not in res_mot.columns: res_mot["Fora"]   = 0
res_mot = res_mot.reset_index()
res_mot["Dentro_h"] = (res_mot["Dentro"]/60).round(2)
res_mot["Fora_h"]   = (res_mot["Fora"]/60).round(2)
res_mot = res_mot.merge(veic_mot, on="Motorista", how="left")

h4 = ["Motorista","Dentro (min)","Dentro (h)","Fora (h)","Veículos Operados"]
for i, h in enumerate(h4, 1):
    ws4.cell(row=4, column=i, value=h)
header_style(ws4, 4, len(h4))

for ri, row in enumerate(res_mot.itertuples(index=False), 5):
    bg = VERDE_CLR if row.Dentro > 0 else (CINZA if ri % 2 == 0 else BRANCO)
    vals = [row.Motorista, round(row.Dentro,1), row.Dentro_h,
            row.Fora_h, row.Veículos_Operados]
    for ci, val in enumerate(vals, 1):
        bold = (row.Dentro > 0 and ci in [1,2,3])
        celula(ws4, ri, ci, val, bold=bold, cor_fundo=bg,
               alinhamento="left" if ci in [1,5] else "center")

larguras4 = [20, 14, 12, 10, 35]
for i, w in enumerate(larguras4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.freeze_panes = "A5"
print("  ✓ Aba 'Resumo Motoristas' criada")

# ── ABA 5: CERCA ELETRÔNICA ───────────────
ws5 = wb.create_sheet("Cerca Eletrônica")
titulo_aba(ws5, "Coordenadas da Cerca Eletrônica",
           "Vértices do polígono que define a Base Operacional")

cerca = pd.read_excel("dados/coordenadas_cerca.xlsx")
h5 = ["Ponto", "Latitude", "Longitude"]
for i, h in enumerate(h5, 1):
    ws5.cell(row=4, column=i, value=h)
header_style(ws5, 4, 3)

for ri, row in enumerate(cerca.itertuples(index=False), 5):
    bg = AZUL_CLR if ri % 2 == 0 else BRANCO
    celula(ws5, ri, 1, row.Ponto, cor_fundo=bg, alinhamento="center")
    celula(ws5, ri, 2, row.Latitude, cor_fundo=bg, alinhamento="center")
    celula(ws5, ri, 3, row.Longitude, cor_fundo=bg, alinhamento="center")

for i, w in enumerate([10, 16, 16], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
print("  ✓ Aba 'Cerca Eletrônica' criada")

# ─────────────────────────────────────────
# 4. SALVAR
# ─────────────────────────────────────────
caminho = "output/base_consolidada_staff_telemetria.xlsx"
wb.save(caminho)
print(f"\n✅ Excel salvo: {caminho}")